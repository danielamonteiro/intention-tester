import os.path
import os
from os import path

import openpyxl
from openpyxl import utils
from openpyxl.styles import Font
from openpyxl.chart import PieChart, Reference

def create_file():
    name_file = input("Insira o nome do arquivo que será gerado com os resultados dos testes: ")
    actual_path = os.path.abspath(os.path.dirname(__file__))
    result_files_path = os.path.join(actual_path, f"../files/result_files/{name_file}.xlsx")
    while path.exists(result_files_path) == True:
        name_file = input(f"O arquivo '{name_file}' já existe. Escolha outro nome: ")
        result_files_path = os.path.join(actual_path, f"../files/result_files/{name_file}.xlsx")
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resultados"
        wb.save(result_files_path)
        print(f"Arquivo '{name_file}.xlsx' criado com sucesso!")
    except:
        print(f"[ATENÇÃO] Erro ao criar o arquivo '{name_file}.xlsx'")
    
    return name_file


def edit_file(name_file, result_list):
    actual_path = os.path.abspath(os.path.dirname(__file__))
    result_file_path = os.path.join(actual_path, f"../files/result_files/{name_file}.xlsx")
    try:
        wb = openpyxl.load_workbook(result_file_path)
        ws = wb.active
        headers = ["Utterance","Intenção Esperada","Intenção Retornada","Confiança","Resultado"]
        ws.append(headers)
        row = ws.row_dimensions[1]
        row.font = Font(bold=True)
        for result in result_list:
            ws.append(result)

        ws.auto_filter.ref = f"A1:E{len(result_list)}"
        ws_result = utils.quote_sheetname(ws.title).replace("'", "")
        wb.save(result_file_path)
        create_chart_file(name_file, ws_result)
    except Exception as e:
        print("[ERRO]", e)

def get_test_files():
    utterances_list = []
    actual_path = os.path.abspath(os.path.dirname(__file__))
    test_files_path = os.path.join(actual_path, f"../files/test_files/")
    
    try:
        for test_files in os.listdir(test_files_path):
            if test_files.endswith(".txt"):
                expected_intent = test_files[:-4]
                actual_file = open(f"{test_files_path}/{test_files}", "r")
                file_to_test = actual_file.read().splitlines()
                actual_file.close()
                for utterance in file_to_test:
                    utterances_list.append([utterance, expected_intent])
        return utterances_list

    except Exception as e:
        print("[ERRO]", e)    
    

def create_chart_file(name_file, ws_result):
    actual_path = os.path.abspath(os.path.dirname(__file__))
    result_file_path = os.path.join(actual_path, f"../files/result_files/{name_file}.xlsx")
    try:
        wb = openpyxl.load_workbook(result_file_path)
        chart_worksheet = wb.create_sheet(title="Gráfico")
        chart_worksheet['A1'] = "Total Sucess"
        chart_worksheet['A2'] = "Total Failed"
        chart_worksheet['B1'] = f'=COUNTIF(${ws_result}.E:E;"Sucess")'
        chart_worksheet['B2'] = f'=COUNTIF(${ws_result}.E:E;"Failed")'

        pie_chart = PieChart()
        values = Reference(chart_worksheet, min_col=1, min_row = 1,
                                            max_col=2, max_row= 2)
        labels = Reference(chart_worksheet, min_col=1, min_row = 1,
                                            max_col=2, max_row= 2)
        pie_chart.add_data(values, titles_from_data = True)
        pie_chart.set_categories(labels) 
        chart_worksheet.add_chart(pie_chart, "A1")
        wb.save(result_file_path)
    except Exception as e:
        print(e)