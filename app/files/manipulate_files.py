import os.path
import os
from os import path

import openpyxl
from openpyxl.utils import get_column_letter

def create_file():
    name_file = input("Insira o nome do arquivo que será gerado com os resultados dos testes: ")
    actual_path = os.path.abspath(os.path.dirname(__file__))
    result_files_path = os.path.join(actual_path, f"../files/result_files/{name_file}.xlsx")
    while path.exists(result_files_path) == True:
        name_file = input(f"O arquivo '{name_file}' já existe. Escolha outro nome: ")
        result_files_path = os.path.join(actual_path, f"../files/result_files/{name_file}.xlsx")
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resultados"
        wb.save(result_files_path)
        print(f"Arquivo '{name_file}.xlsx' criado com sucesso!")
    except:
        print(f"[ATENÇÃO] Erro ao criar o arquivo '{name_file}.xlsx'")
    
    return name_file


def edit_file(name_file, result_list):
    actual_path = os.path.abspath(os.path.dirname(__file__))
    result_file_path = os.path.join(actual_path, f"../files/result_files/{name_file}.xlsx")
    wb = openpyxl.load_workbook(result_file_path)
    ws = wb.active
    headers = ["Utterance","Intenção Esperada","Intenção Retornada","Confiança","Resultado"]
    ws.append(headers)
    for result in result_list:
        ws.append(result)

    wb.save(result_file_path)

def get_test_files():
    utterances_list = []
    actual_path = os.path.abspath(os.path.dirname(__file__))
    test_files_path = os.path.join(actual_path, f"../files/test_files/")
    
    for test_files in os.listdir(test_files_path):
        expected_intent = test_files[:-4]
        actual_file = open(f"{test_files_path}/{test_files}", "r")
        file_to_test = actual_file.read().splitlines()
        actual_file.close()
        for utterance in file_to_test:
            utterances_list.append([utterance, expected_intent])
    
    return utterances_list