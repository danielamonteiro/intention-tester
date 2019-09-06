import os.path

import openpyxl
from openpyxl.utils import get_column_letter

def create_file():
    name_file = input("Insira o nome do arquivo que será gerado com os resulados dos testes: ")
    my_path = os.path.abspath(os.path.dirname(__file__))
    path = os.path.join(my_path, f"../files/result_files/{name_file}.xlsx")
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resultados"
        wb.save(path)
        print(f"Arquivo '{name_file}.xlsx' criado com sucesso!")
    except:
        print(f"[ATENÇÃO] Erro ao criar o arquivo '{name_file}.xlsx'")
    
    return name_file


def see_files_infos(name_file):
    try:
        my_path = os.path.abspath(os.path.dirname(__file__))
        path = os.path.join(my_path, f"../files/result_files/{name_file}.xlsx")
        workbook_file = openpyxl.load_workbook(path)
        print(workbook_file.sheetnames)
    except:
        print(f"[ATENÇÃO] Erro ao abrir o arquivo '{name_file}.xlsx' verifique se ele existe na pasta result_files!")


def edit_file(name_file, message):
    my_path = os.path.abspath(os.path.dirname(__file__))
    path = os.path.join(my_path, f"../files/result_files/{name_file}.xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    Row_size = len(message)

    for column, text in enumerate(message, start=1):
        ws.cell(column=column, row=2, value=text)

    wb.save(path)